#!/usr/bin/env python3
"""
USPTO Opposition Trademark Scraper - Web Application
Web interface for retrieving US and International classes from opposition pleaded applications.
"""

import streamlit as st
import requests
from bs4 import BeautifulSoup
import json
import pandas as pd
from typing import List, Dict
import time
import io
import base64
import os
from datetime import datetime
import re
import anthropic
from dotenv import load_dotenv
from database_cache import TrademarkCache
from research_format_exporter import create_research_format_excel, create_copyable_research_format
from bulk_copyable_formatter import create_bulk_copyable_format, create_bulk_copyable_excel

# Load environment variables from .env file
load_dotenv()


class USPTOOppositionScraper:
    """Scraper for USPTO opposition trademark data."""

    def __init__(self, api_key: str, claude_vision_api_key: str = None, anthropic_api_key: str = None):
        """Initialize scraper with API key."""
        self.api_key = api_key
        self.claude_vision_api_key = claude_vision_api_key
        self.anthropic_api_key = anthropic_api_key
        self.tsdr_base_url = "https://tsdrapi.uspto.gov/ts/cd/casestatus/sn{}/info.json"
        self.tsdr_image_url = "https://tsdrapi.uspto.gov/ts/cd/rawImage/{}"
        self.ttabvue_base_url = "https://ttabvue.uspto.gov/ttabvue/v"
        self.session = requests.Session()
        self.session.headers.update({'USPTO-API-KEY': self.api_key})

        # Initialize database cache
        self.cache = TrademarkCache(cache_ttl_days=30)
        self.session_stats = {'cache_hits': 0, 'cache_misses': 0, 'api_calls_saved': 0}

    def get_serial_numbers_from_opposition(self, opposition_number: str, proceeding_type: str = 'OPP') -> List[Dict[str, str]]:
        """Scrape TTABVue to get serial numbers from pleaded applications section."""

        params = {
            'pno': opposition_number,
            'pty': proceeding_type
        }

        try:
            response = requests.get(self.ttabvue_base_url, params=params, timeout=30)
            response.raise_for_status()
        except requests.RequestException as e:
            st.error(f"Error fetching TTABVue page: {e}")
            return []

        soup = BeautifulSoup(response.text, 'html.parser')
        serial_numbers = []
        import re

        # Find the "Pleaded applications and registrations" heading in table
        # Look specifically for the TH with class t3 containing this text
        pleaded_found = False
        pleaded_row_index = -1
        current_table = None

        for table in soup.find_all('table'):
            rows = table.find_all('tr')
            for idx, row in enumerate(rows):
                # Look specifically for TH with class t3 (the section header style)
                header_cell = row.find('th', class_='t3')
                if header_cell and 'pleaded applications and registrations' in header_cell.get_text().lower():
                    pleaded_found = True
                    pleaded_row_index = idx
                    current_table = table
                    break
                if pleaded_found:
                    break
            if pleaded_found:
                break

        if current_table and pleaded_row_index >= 0:
            # Extract serial numbers ONLY after the pleaded section heading
            rows = current_table.find_all('tr')

            # Find the end marker - look for major section breaks AFTER pleaded section
            # We want to stop at "Prosecution History" or similar major sections
            end_row_index = len(rows)
            for idx in range(pleaded_row_index + 1, len(rows)):
                row = rows[idx]
                row_text = row.get_text().strip().lower()

                # Check if this is a major section header (has class t2b AND has actual text content)
                section_cell = row.find('td', class_='t2b')
                if section_cell and section_cell.get_text().strip():
                    # This is a major section break with content
                    end_row_index = idx
                    break

                # Also stop at Prosecution History
                if 'prosecution history' in row_text:
                    end_row_index = idx
                    break

            # Now extract serial numbers only between pleaded section and end marker
            i = pleaded_row_index + 1
            while i < end_row_index:
                row = rows[i]
                # Look for "Serial #:" in the row
                for cell in row.find_all('th'):
                    if 'Serial #:' in cell.get_text():
                        # Next cell should contain the serial number
                        serial_link = row.find('a', href=lambda x: x and 'tsdr.uspto.gov' in x and 'caseNumber=' in x)
                        if serial_link:
                            serial_match = re.search(r'\d{8}', serial_link.get_text())
                            if serial_match:
                                sn = serial_match.group(0)

                                # Find mark name in subsequent rows
                                mark_name = 'Unknown'
                                # Look ahead for "Mark:" row
                                for j in range(i+1, min(i+5, end_row_index)):
                                    next_row = rows[j]
                                    for next_cell in next_row.find_all('th'):
                                        if 'Mark:' in next_cell.get_text():
                                            # Mark value is in adjacent td
                                            mark_td = next_row.find('td')
                                            if mark_td:
                                                mark_name = mark_td.get_text(strip=True)
                                            break
                                    if mark_name != 'Unknown':
                                        break

                                serial_numbers.append({
                                    'serial_number': sn,
                                    'mark_name': mark_name
                                })
                                break
                i += 1

        # Remove duplicates while preserving order
        seen = set()
        unique_serials = []
        for item in serial_numbers:
            sn = item['serial_number']
            if sn not in seen:
                seen.add(sn)
                unique_serials.append(item)

        return unique_serials

    def get_opposition_dates(self, opposition_number: str, proceeding_type: str = 'OPP') -> Dict[str, str]:
        """
        Extract filing date and termination/last date from opposition page.
        Returns: {'filing_date': 'MM/DD/YYYY', 'termination_date': 'MM/DD/YYYY'}
        """
        params = {
            'pno': opposition_number,
            'pty': proceeding_type
        }

        try:
            response = requests.get(self.ttabvue_base_url, params=params, timeout=30)
            response.raise_for_status()
        except requests.RequestException:
            return {'filing_date': None, 'termination_date': None}

        soup = BeautifulSoup(response.text, 'html.parser')
        filing_date = None
        termination_date = None

        # Extract filing date from the summary section
        # Look for "Filing Date:" label
        for row in soup.find_all('tr'):
            cells = row.find_all(['th', 'td'])
            for i, cell in enumerate(cells):
                cell_text = cell.get_text().strip()
                if 'Filing Date:' in cell_text or 'filing date' in cell_text.lower():
                    # Next cell should contain the date
                    if i + 1 < len(cells):
                        date_text = cells[i + 1].get_text().strip()
                        date_match = re.search(r'\b(\d{2}/\d{2}/\d{4})\b', date_text)
                        if date_match:
                            filing_date = date_match.group(1)
                            break

        # Extract last/termination date from Prosecution History section
        # Find the Prosecution History table
        prosecution_history_found = False
        for table in soup.find_all('table'):
            rows = table.find_all('tr')
            for row in rows:
                # Look for "Prosecution History" header
                if 'prosecution history' in row.get_text().lower():
                    prosecution_history_found = True
                    break

            if prosecution_history_found:
                # Find all rows with dates in this table
                # The last row with a date is the termination/last action date
                dates_found = []
                for row in rows:
                    row_text = row.get_text()
                    # Look for date pattern MM/DD/YYYY
                    date_matches = re.findall(r'\b(\d{2}/\d{2}/\d{4})\b', row_text)
                    if date_matches:
                        # Take the first date from each row (usually the action date)
                        dates_found.append(date_matches[0])

                # The last date in the prosecution history is the termination/last action date
                if dates_found:
                    termination_date = dates_found[-1]
                break

        return {
            'filing_date': filing_date,
            'termination_date': termination_date
        }

    def get_party_info(self, opposition_number: str, proceeding_type: str = 'OPP') -> Dict:
        """
        Extract plaintiff and defendant information from opposition page.
        Returns: {
            'plaintiff_name': str,
            'defendant_name': str,
            'plaintiff_serials': [{serial_number, mark_name}, ...],
            'defendant_serials': [{serial_number, mark_name}, ...]
        }
        """
        params = {
            'pno': opposition_number,
            'pty': proceeding_type
        }

        try:
            response = requests.get(self.ttabvue_base_url, params=params, timeout=30)
            response.raise_for_status()
        except requests.RequestException:
            return {
                'plaintiff_name': None,
                'defendant_name': None,
                'plaintiff_serials': [],
                'defendant_serials': []
            }

        soup = BeautifulSoup(response.text, 'html.parser')

        plaintiff_name = None
        defendant_name = None
        plaintiff_serials = []
        defendant_serials = []

        # Find Plaintiff and Defendant sections
        current_party = None

        for table in soup.find_all('table'):
            rows = table.find_all('tr')

            for row in rows:
                # Check for Plaintiff or Defendant section header
                section_cell = row.find('td', class_='t2b')
                if section_cell:
                    section_text = section_cell.get_text().strip()
                    if section_text == 'Plaintiff':
                        current_party = 'plaintiff'
                        continue
                    elif section_text == 'Defendant':
                        current_party = 'defendant'
                        continue
                    elif section_text and current_party:
                        # New major section, reset
                        current_party = None

                # Extract party name
                if current_party:
                    name_cell = row.find('th', class_='t3', string=lambda x: x and 'Name:' in x)
                    if name_cell:
                        name_link = row.find('a', href=lambda x: x and 'pnam=' in x)
                        if name_link:
                            name = name_link.get_text().strip()
                            if current_party == 'plaintiff' and not plaintiff_name:
                                plaintiff_name = name
                            elif current_party == 'defendant' and not defendant_name:
                                defendant_name = name

        # Now extract serial numbers and associate with plaintiff/defendant
        # Find the "Pleaded applications and registrations" section
        pleaded_found = False
        pleaded_row_index = -1
        current_table = None

        for table in soup.find_all('table'):
            rows = table.find_all('tr')
            for idx, row in enumerate(rows):
                header_cell = row.find('th', class_='t3')
                if header_cell and 'pleaded applications and registrations' in header_cell.get_text().lower():
                    pleaded_found = True
                    pleaded_row_index = idx
                    current_table = table
                    break
            if pleaded_found:
                break

        if current_table and pleaded_row_index >= 0:
            rows = current_table.find_all('tr')

            # Find end of pleaded section
            end_row_index = len(rows)
            for idx in range(pleaded_row_index + 1, len(rows)):
                row = rows[idx]
                section_cell = row.find('td', class_='t2b')
                if section_cell and section_cell.get_text().strip():
                    end_row_index = idx
                    break
                if 'prosecution history' in row.get_text().lower():
                    end_row_index = idx
                    break

            # Extract serials and determine ownership
            current_owner = None
            i = pleaded_row_index + 1

            while i < end_row_index:
                row = rows[i]

                # Check for "Owned by" row
                owned_by_cell = row.find('th', string=lambda x: x and 'Owned by:' in x)
                if owned_by_cell:
                    owner_td = row.find('td')
                    if owner_td:
                        owner_name = owner_td.get_text().strip()
                        # Match with plaintiff or defendant
                        if plaintiff_name and plaintiff_name.lower() in owner_name.lower():
                            current_owner = 'plaintiff'
                        elif defendant_name and defendant_name.lower() in owner_name.lower():
                            current_owner = 'defendant'

                # Check for Serial # row
                for cell in row.find_all('th'):
                    if 'Serial #:' in cell.get_text():
                        serial_link = row.find('a', href=lambda x: x and 'tsdr.uspto.gov' in x and 'caseNumber=' in x)
                        if serial_link:
                            serial_match = re.search(r'\d{8}', serial_link.get_text())
                            if serial_match:
                                sn = serial_match.group(0)

                                # Find mark name
                                mark_name = 'Unknown'
                                for j in range(i+1, min(i+5, end_row_index)):
                                    next_row = rows[j]
                                    for next_cell in next_row.find_all('th'):
                                        if 'Mark:' in next_cell.get_text():
                                            mark_td = next_row.find('td')
                                            if mark_td:
                                                mark_name = mark_td.get_text(strip=True)
                                            break
                                    if mark_name != 'Unknown':
                                        break

                                serial_info = {
                                    'serial_number': sn,
                                    'mark_name': mark_name
                                }

                                if current_owner == 'plaintiff':
                                    plaintiff_serials.append(serial_info)
                                elif current_owner == 'defendant':
                                    defendant_serials.append(serial_info)
                                break
                i += 1

        return {
            'plaintiff_name': plaintiff_name,
            'defendant_name': defendant_name,
            'plaintiff_serials': plaintiff_serials,
            'defendant_serials': defendant_serials
        }

    def get_opposition_result(self, opposition_number: str, proceeding_type: str = 'OPP') -> Dict:
        """
        Extract filing date, termination date, and result from opposition page.
        Returns: {
            'filing_date': 'MM/DD/YYYY',
            'termination_date': 'MM/DD/YYYY',
            'result': 0 or 1 (0=Dismissed, 1=Sustained)
        }
        """
        params = {
            'pno': opposition_number,
            'pty': proceeding_type
        }

        try:
            response = requests.get(self.ttabvue_base_url, params=params, timeout=30)
            response.raise_for_status()
        except requests.RequestException:
            return {'filing_date': None, 'termination_date': None, 'result': None}

        soup = BeautifulSoup(response.text, 'html.parser')
        filing_date = None
        termination_date = None
        result = None

        # Find Prosecution History section
        prosecution_history_found = False
        for table in soup.find_all('table'):
            rows = table.find_all('tr')
            for row in rows:
                if 'prosecution history' in row.get_text().lower():
                    prosecution_history_found = True
                    break

            if prosecution_history_found:
                # Look for FILED AND FEE and TERMINATED rows
                for row in rows:
                    row_text = row.get_text()

                    # Find filing date
                    if 'FILED AND FEE' in row_text:
                        cells = row.find_all('td')
                        if len(cells) >= 2:
                            date_match = re.search(r'\b(\d{2}/\d{2}/\d{4})\b', cells[1].get_text())
                            if date_match:
                                filing_date = date_match.group(1)

                    # Find termination date and result
                    if 'TERMINATED' in row_text:
                        cells = row.find_all('td')
                        if len(cells) >= 2:
                            date_match = re.search(r'\b(\d{2}/\d{2}/\d{4})\b', cells[1].get_text())
                            if date_match:
                                termination_date = date_match.group(1)

                    # Check for result (Dismissed or Sustained)
                    if 'SUSTAINED' in row_text.upper():
                        result = 1
                    elif 'DISMISSED' in row_text.upper():
                        result = 0

                break

        return {
            'filing_date': filing_date,
            'termination_date': termination_date,
            'result': result
        }

    def analyze_opposition_complete(self, opposition_number: str, company_name: str, proceeding_type: str = 'OPP', progress_callback=None) -> Dict:
        """
        Complete opposition analysis matching the Excel format requirements.
        Returns one row of data per opposition with all required fields.
        """
        if progress_callback:
            progress_callback(0, f"Analyzing opposition {opposition_number}...")

        # Get party information
        party_info = self.get_party_info(opposition_number, proceeding_type)

        # Determine if company is plaintiff or defendant
        is_plaintiff = 0
        alt_name = None

        if party_info['plaintiff_name'] and company_name.lower() in party_info['plaintiff_name'].lower():
            is_plaintiff = 1
            alt_name = party_info['plaintiff_name']
        elif party_info['defendant_name'] and company_name.lower() in party_info['defendant_name'].lower():
            is_plaintiff = 0
            alt_name = party_info['defendant_name']

        # Get plaintiff's marks only
        plaintiff_serials = party_info['plaintiff_serials']
        marks_count = len(plaintiff_serials)

        if progress_callback:
            progress_callback(0.3, f"Processing {marks_count} plaintiff marks...")

        # Get classes and mark types for plaintiff marks
        unique_us_classes = set()
        unique_int_classes = set()
        tm_type_counts = {1: 0, 2: 0, 3: 0}  # Standard, Stylized, Slogan
        mark_details = []

        for idx, serial_info in enumerate(plaintiff_serials):
            if progress_callback:
                progress_callback(0.3 + (0.4 * (idx + 1) / len(plaintiff_serials)),
                                f"Processing mark {idx+1}/{marks_count}")

            sn = serial_info['serial_number']
            mark_name = serial_info['mark_name']

            # Get classes
            class_data = self.get_classes_from_serial(sn)

            # Add to unique sets
            for uc in class_data['us_classes']:
                unique_us_classes.add(uc['code'])
            for ic in class_data['international_classes']:
                unique_int_classes.add(ic['code'])

            # Count mark type
            mark_type = class_data.get('mark_type', 0)
            if mark_type in tm_type_counts:
                tm_type_counts[mark_type] += 1

            mark_details.append({
                'serial_number': sn,
                'mark_name': mark_name,
                'mark_type': mark_type
            })

            time.sleep(0.3)

        if progress_callback:
            progress_callback(0.8, "Extracting dates and result...")

        # Get dates and result
        result_info = self.get_opposition_result(opposition_number, proceeding_type)

        if progress_callback:
            progress_callback(1.0, "Complete!")

        return {
            'opposition_number': opposition_number,
            'alt_name': alt_name,
            'plaintiff': is_plaintiff,
            'marks': marks_count,
            'us_gs': len(unique_us_classes),
            'int_gs': len(unique_int_classes),
            'opp_start_date': result_info['filing_date'],
            'opp_end_date': result_info['termination_date'],
            'result': result_info['result'],
            'tm_type_1': tm_type_counts[1],
            'tm_type_2': tm_type_counts[2],
            'tm_type_3': tm_type_counts[3],
            'mark_details': mark_details  # For Excel serial number columns
        }

    def is_text_a_slogan(self, text: str, anthropic_api_key: str) -> bool:
        """Use Claude API to determine if text is a marketing slogan with retry logic."""
        import anthropic

        client = anthropic.Anthropic(api_key=anthropic_api_key)
        max_retries = 3

        for attempt in range(max_retries):
            try:
                # Add delay between attempts
                if attempt > 0:
                    wait_time = min(2 ** attempt * 2, 30)  # 2s, 4s, 8s (max 30s)
                    time.sleep(wait_time)
                else:
                    time.sleep(0.3)  # Small delay on first attempt

                message = client.messages.create(
                    model="claude-sonnet-4-5-20250929",
                    max_tokens=50,
                    messages=[
                        {
                            "role": "user",
                            "content": f"""Analyze this text and determine if it's a marketing slogan or just plain text/brand name.

Text: "{text}"

Reply with ONLY "YES" if it's a marketing slogan (a catchy phrase used for marketing/advertising), or "NO" if it's just a plain text, brand name, or descriptive phrase."""
                        }
                    ]
                )

                response_text = message.content[0].text.strip().upper()
                return response_text == "YES"

            except anthropic.InternalServerError as e:
                if attempt < max_retries - 1:
                    print(f"  -> Slogan detection API overloaded, retrying... ({attempt + 1}/{max_retries})")
                    continue
                else:
                    print(f"  -> Slogan detection failed after retries, using fallback")
                    return False

            except anthropic.RateLimitError as e:
                if attempt < max_retries - 1:
                    print(f"  -> Slogan detection rate limited, retrying... ({attempt + 1}/{max_retries})")
                    continue
                else:
                    return False

            except Exception as e:
                # Any other error, fall back
                print(f"  -> Slogan detection error: {str(e)}")
                return False

        # If all retries exhausted
        return False

    def classify_mark_image(self, serial_number: str, anthropic_api_key: str = None) -> int:
        """
        Classify trademark image using Claude Vision API.
        Returns: 0 (no image), 1 (standard text), 2 (stylized/design), 3 (slogan)
        Note: Returns 0 ONLY if image contains "No Image exists" text
        """
        if not anthropic_api_key:
            # Default to Type 2 if no API key provided
            return 2

        # Download image
        image_url = self.tsdr_image_url.format(serial_number)
        try:
            response = self.session.get(image_url, timeout=30)
            response.raise_for_status()
            image_content = response.content
        except requests.RequestException:
            # Default to Type 2 on download failure
            return 2

        # Determine media type by checking file signatures (magic bytes)
        image_media_type = "image/jpeg"  # default
        is_tiff = False

        if image_content[:8] == b'\x89PNG\r\n\x1a\n':
            image_media_type = "image/png"
        elif image_content[:6] in (b'GIF87a', b'GIF89a'):
            image_media_type = "image/gif"
        elif image_content[:4] == b'\xff\xd8\xff\xe0' or image_content[:4] == b'\xff\xd8\xff\xe1':
            image_media_type = "image/jpeg"
        elif image_content[:2] == b'\xff\xd8':
            image_media_type = "image/jpeg"
        elif image_content[:4] == b'RIFF' and image_content[8:12] == b'WEBP':
            image_media_type = "image/webp"
        elif image_content[:4] in (b'II*\x00', b'MM\x00*'):  # TIFF (little-endian and big-endian)
            is_tiff = True

        # Convert TIFF to JPEG (Claude doesn't support TIFF)
        if is_tiff:
            try:
                from PIL import Image
                import io
                img = Image.open(io.BytesIO(image_content))
                # Convert to RGB if necessary
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                # Save as JPEG
                jpeg_buffer = io.BytesIO()
                img.save(jpeg_buffer, format='JPEG', quality=95)
                image_content = jpeg_buffer.getvalue()
                image_media_type = "image/jpeg"
            except Exception as e:
                print(f"Error converting TIFF to JPEG: {str(e)}")
                return 0

        # Encode image to base64
        image_base64 = base64.b64encode(image_content).decode('utf-8')

        # Use Claude Vision API with retry logic and rate limiting
        import anthropic
        client = anthropic.Anthropic(api_key=anthropic_api_key)

        max_retries = 5
        message = None

        for attempt in range(max_retries):
            try:
                # Add delay between attempts to avoid rate limiting
                if attempt > 0:
                    wait_time = min(2 ** attempt * 3, 60)  # Exponential backoff: 3s, 6s, 12s, 24s, 48s (max 60s)
                    print(f"  -> Retry {attempt}/{max_retries} for serial {serial_number}, waiting {wait_time}s...")
                    time.sleep(wait_time)

                # Add small delay even on first attempt to avoid hammering the API
                if attempt == 0:
                    time.sleep(0.5)  # 500ms rate limiting between calls

                message = client.messages.create(
                    model="claude-sonnet-4-5-20250929",
                    max_tokens=1024,
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "image",
                                    "source": {
                                        "type": "base64",
                                        "media_type": image_media_type,
                                        "data": image_base64,
                                    },
                                },
                                {
                                    "type": "text",
                                    "text": """Analyze this trademark image and provide:
1. All text detected in the image (word for word)
2. Whether there are any logos, symbols, or graphic design elements
3. Visual characteristics: font styling, colors, decorative elements, shapes, patterns
4. Overall visual complexity (simple/moderate/complex)

Format your response as:
TEXT: [all text found]
HAS_LOGO: [yes/no]
HAS_DESIGN: [yes/no]
VISUAL_ELEMENTS: [list of visual characteristics]
COMPLEXITY: [simple/moderate/complex]"""
                                }
                            ],
                        }
                    ],
                )

                # Success! Break out of retry loop
                break

            except anthropic.InternalServerError as e:
                # Handle overloaded/500 errors
                if attempt < max_retries - 1:
                    print(f"  -> Anthropic API overloaded (500), retrying... ({attempt + 1}/{max_retries})")
                    continue
                else:
                    print(f"  -> Anthropic API overloaded after {max_retries} retries. Using fallback classification.")
                    # Fallback: Return Type 2 (Stylized) as default
                    return 2

            except anthropic.RateLimitError as e:
                # Handle rate limit errors
                if attempt < max_retries - 1:
                    print(f"  -> Rate limit exceeded, retrying... ({attempt + 1}/{max_retries})")
                    continue
                else:
                    print(f"  -> Rate limit exceeded after {max_retries} retries. Using fallback classification.")
                    return 2

            except anthropic.APIError as e:
                # Handle other Anthropic API errors
                if attempt < max_retries - 1:
                    print(f"  -> Anthropic API error: {str(e)}, retrying... ({attempt + 1}/{max_retries})")
                    continue
                else:
                    print(f"  -> Anthropic API error after {max_retries} retries. Using fallback classification.")
                    return 2

        # If we exhausted all retries without success
        if message is None:
            print(f"  -> All retries exhausted for serial {serial_number}. Using fallback classification.")
            return 2

        try:

            # Parse Claude's response
            response_text = message.content[0].text

            # Extract information from response
            detected_text = ""
            has_logo = False
            has_design = False
            labels = []

            for line in response_text.split('\n'):
                if line.startswith('TEXT:'):
                    detected_text = line.replace('TEXT:', '').strip()
                elif line.startswith('HAS_LOGO:'):
                    has_logo = 'yes' in line.lower()
                elif line.startswith('HAS_DESIGN:'):
                    has_design = 'yes' in line.lower()
                elif line.startswith('VISUAL_ELEMENTS:'):
                    elements = line.replace('VISUAL_ELEMENTS:', '').strip().lower()
                    # Create pseudo-labels from visual elements
                    if elements:
                        for elem in elements.split(','):
                            labels.append((elem.strip(), 0.8))
                elif line.startswith('COMPLEXITY:'):
                    complexity = line.replace('COMPLEXITY:', '').strip().lower()
                    if 'complex' in complexity or 'moderate' in complexity:
                        labels.append(('complex', 0.9))
                        labels.append(('design', 0.9))

            # CRITICAL: Check if this is the "No Image exists" placeholder
            # Only return 0 (No Image) if the image contains this specific text
            print(f"  -> Serial {serial_number}: Detected text = '{detected_text[:200]}'")
            print(f"  -> Serial {serial_number}: Word count = {len(words) if 'words' in locals() else 0}")
            print(f"  -> Serial {serial_number}: Has logo = {has_logo}, Has design = {has_design}")

            if detected_text and 'no image exists' in detected_text.lower():
                print(f"  -> Serial {serial_number}: Detected 'No Image exists' placeholder → returning 0")
                return 0

            # Classification logic - count only alphanumeric words, ignore symbols
            import re
            # Extract only words that contain letters or numbers (ignore pure symbols)
            words = re.findall(r'\b[a-zA-Z0-9]+\b', detected_text) if detected_text else []
            word_count = len(words)

            # Rule 1: No text detected (pure logo/symbol/design) → Type 2
            if not detected_text or word_count == 0:
                return 2

            # Rule 2: ANY logo detection → Type 2
            if has_logo:
                return 2

            # Rule 3: Check for ANY visual indicators (be very aggressive)
            # If there are ANY labels at all with reasonable confidence, likely has visual elements
            if len(labels) >= 3:  # Just having multiple labels suggests visual complexity
                return 2

            # Check for ANY design/visual keywords at very low threshold
            design_style_keywords = [
                'art', 'illustration', 'drawing', 'painting', 'artwork', 'graphics', 'design',
                'creative', 'logo', 'symbol', 'icon', 'emblem', 'badge', 'insignia',
                'font', 'calligraphy', 'typography', 'ornate', 'decorative', 'stylized',
                'artistic', 'handwriting', 'script', 'cursive', 'fancy', 'vintage',
                'modern', 'retro', 'bold', 'italic', 'visual', 'graphic', 'rectangle',
                'pattern', 'shape', 'circle', 'square', 'line', 'color', 'black', 'white'
            ]
            # Lower threshold to 0.3 to catch more design elements
            has_any_styling = any(
                keyword in label_text and score > 0.3
                for label_text, score in labels
                for keyword in design_style_keywords
            )

            if has_any_styling:
                return 2

            # Rule 4: Exception for Type 3 (Slogan) - VERY strict criteria
            # Only Type 3 if: 3+ words AND absolutely no visual indicators
            if word_count >= 3:
                # Check if this is truly plain text with no styling
                # Must have very few labels (indicating simple image)
                if len(labels) <= 2:
                    return 3
                else:
                    return 2  # Has labels, so likely has styling

            # Rule 5: Exception for Type 1 (Standard) - EXTREMELY strict
            # Only Type 1 if: 1-2 words AND absolutely minimal visual complexity
            # This should be VERY rare
            if word_count > 0 and word_count <= 2:
                # Must have almost no labels at all
                if len(labels) <= 1:
                    return 1
                else:
                    return 2  # Any labels suggest visual styling

            # Default to Type 2 (Stylized) for anything unclear
            return 2

        except Exception as e:
            # Log the error for debugging
            print(f"Error classifying mark {serial_number}: {str(e)}")
            print(f"Exception type: {type(e).__name__}")
            import traceback
            traceback.print_exc()

            # Add specific handling for common errors
            if "PIL" in str(e) or "Image" in str(e):
                print(f"  -> PIL/Image error detected. Image may be corrupted or in unsupported format.")
            if "anthropic" in str(e).lower():
                print(f"  -> Anthropic API error. Check API key, quota, or image format.")

            # FALLBACK: Try to extract text using basic OCR as a last resort
            print(f"  -> Attempting fallback classification using OCR...")
            try:
                from PIL import Image
                import io
                import pytesseract

                # Try OCR on the image
                img = Image.open(io.BytesIO(image_content))
                extracted_text = pytesseract.image_to_string(img).strip()

                print(f"  -> OCR extracted text: {extracted_text[:100]}")

                # Check for "No Image exists" message
                if 'no image exists' in extracted_text.lower():
                    return 0

                # Count words
                import re
                words = re.findall(r'\b[a-zA-Z0-9]+\b', extracted_text) if extracted_text else []
                word_count = len(words)

                print(f"  -> Word count from OCR: {word_count}")

                # Simple classification based on word count
                if word_count == 0:
                    # No text detected, likely a design/logo
                    return 2
                elif word_count >= 3:
                    # 3+ words = slogan
                    return 3
                elif word_count <= 2:
                    # 1-2 words = standard text (conservative)
                    return 1
                else:
                    return 2

            except Exception as ocr_error:
                print(f"  -> OCR fallback also failed: {str(ocr_error)}")
                # Ultimate fallback: Type 2 (most common)
                print(f"  -> Defaulting to Type 2 (Stylized/Design) as last resort")
                return 2

    def get_classes_from_serial(self, serial_number: str) -> Dict:
        """Fetch US and International classes for a serial number via TSDR API.

        Implements retry logic with exponential backoff for reliability.
        CACHE-FIRST: Check database before making API calls.
        """
        import time

        # STEP 1: Check cache first
        cached_data = self.cache.get_cached_data(serial_number)
        if cached_data:
            # Cache hit! Return immediately
            self.session_stats['cache_hits'] += 1
            self.session_stats['api_calls_saved'] += 1
            self.cache.increment_api_savings('tsdr')
            if cached_data.get('mark_type', 0) != 0:
                self.cache.increment_api_savings('anthropic')

            return {
                'us_classes': cached_data['us_classes'],
                'international_classes': cached_data['international_classes'],
                'description': cached_data['description'],
                'mark_type': cached_data.get('mark_type', 0),
                'filing_date': cached_data.get('filing_date', ''),
                'from_cache': True
            }
        else:
            self.session_stats['cache_misses'] += 1

        # STEP 2: Cache miss - fetch from API
        url = self.tsdr_base_url.format(serial_number)
        max_retries = 3
        base_delay = 1  # seconds

        # Try fetching data with retries
        for attempt in range(max_retries):
            try:
                # Increased timeout for slow API responses
                response = self.session.get(url, timeout=60)
                response.raise_for_status()
                data = response.json()

                # Success! Break out of retry loop
                if attempt > 0:
                    print(f"✓ Serial {serial_number}: Successfully fetched on attempt {attempt + 1}")
                break

            except requests.Timeout as e:
                if attempt < max_retries - 1:
                    delay = base_delay * (2 ** attempt)  # Exponential backoff: 1s, 2s, 4s
                    print(f"⚠ Serial {serial_number}: Timeout on attempt {attempt + 1}/{max_retries}. Retrying in {delay}s...")
                    time.sleep(delay)
                else:
                    print(f"✗ Serial {serial_number}: Failed after {max_retries} attempts - Timeout: {str(e)}")
                    return {
                        'us_classes': [],
                        'international_classes': [],
                        'description': '',
                        'mark_type': 0,
                        'error': f'Timeout after {max_retries} attempts'
                    }

            except requests.HTTPError as e:
                if attempt < max_retries - 1 and e.response.status_code in [500, 502, 503, 504]:
                    # Retry on server errors
                    delay = base_delay * (2 ** attempt)
                    print(f"⚠ Serial {serial_number}: HTTP {e.response.status_code} on attempt {attempt + 1}/{max_retries}. Retrying in {delay}s...")
                    time.sleep(delay)
                else:
                    # Don't retry on client errors (404, 400, etc.) or after max retries
                    print(f"✗ Serial {serial_number}: HTTP Error {e.response.status_code}: {str(e)}")
                    return {
                        'us_classes': [],
                        'international_classes': [],
                        'description': '',
                        'mark_type': 0,
                        'error': f'HTTP {e.response.status_code}'
                    }

            except requests.RequestException as e:
                if attempt < max_retries - 1:
                    delay = base_delay * (2 ** attempt)
                    print(f"⚠ Serial {serial_number}: Request error on attempt {attempt + 1}/{max_retries}. Retrying in {delay}s...")
                    time.sleep(delay)
                else:
                    print(f"✗ Serial {serial_number}: Failed after {max_retries} attempts - {type(e).__name__}: {str(e)}")
                    return {
                        'us_classes': [],
                        'international_classes': [],
                        'description': '',
                        'mark_type': 0,
                        'error': f'{type(e).__name__}'
                    }

        try:
            trademark = data['trademarks'][0]
            gs_list = trademark.get('gsList', [])

            # Get filing date
            status = trademark.get('status', {})
            filing_date = status.get('filingDate', '')

            us_classes = []
            international_classes = []
            descriptions = []

            # Track seen class codes to remove duplicates
            seen_us_codes = set()
            seen_intl_codes = set()

            for gs in gs_list:
                for uc in gs.get('usClasses', []):
                    code = uc['code']
                    if code not in seen_us_codes:
                        us_classes.append({
                            'code': code,
                            'description': uc['description']
                        })
                        seen_us_codes.add(code)

                for ic in gs.get('internationalClasses', []):
                    code = ic['code']
                    if code not in seen_intl_codes:
                        international_classes.append({
                            'code': code,
                            'description': ic['description']
                        })
                        seen_intl_codes.add(code)

                desc = gs.get('description', '')
                if desc:
                    descriptions.append(desc)

            # Classify mark image
            mark_type = self.classify_mark_image(serial_number, self.anthropic_api_key)

            result = {
                'us_classes': us_classes,
                'international_classes': international_classes,
                'description': ' | '.join(descriptions) if descriptions else '',
                'mark_type': mark_type,
                'filing_date': filing_date
            }

            # STEP 3: Save to cache for future use
            self.cache.save_to_cache(serial_number, result)

            return result

        except (KeyError, IndexError):
            error_result = {
                'us_classes': [],
                'international_classes': [],
                'description': '',
                'mark_type': 0,
                'filing_date': ''
            }
            # Don't cache errors
            return error_result

    def scrape_opposition(self, opposition_number: str, proceeding_type: str = 'OPP', progress_callback=None) -> Dict:
        """Main method to scrape opposition data."""

        if progress_callback:
            progress_callback(0, "Fetching serial numbers from TTABVue...")

        serials = self.get_serial_numbers_from_opposition(opposition_number, proceeding_type)

        if not serials:
            return {
                'opposition_number': opposition_number,
                'serial_count': 0,
                'data': [],
                'unique_us_classes': [],
                'unique_international_classes': [],
                'filing_date': '',
                'termination_date': '',
                'result': None
            }

        # Get opposition filing and termination dates
        result_info = self.get_opposition_result(opposition_number, proceeding_type)

        all_data = []
        unique_us_classes = set()
        unique_international_classes = set()
        total_us_classes = 0
        total_international_classes = 0

        total = len(serials)
        failed_serials = []  # Track failed serial numbers

        for idx, serial_info in enumerate(serials, 1):
            if progress_callback:
                progress_callback(idx / total, f"Processing {idx}/{total}: {serial_info['serial_number']}")

            sn = serial_info['serial_number']
            mark_name = serial_info['mark_name']

            class_data = self.get_classes_from_serial(sn)

            # Track failures for error reporting
            if 'error' in class_data:
                failed_serials.append({
                    'serial_number': sn,
                    'mark_name': mark_name,
                    'error': class_data['error']
                })

            us_codes = [c['code'] for c in class_data['us_classes']]
            intl_codes = [c['code'] for c in class_data['international_classes']]

            unique_us_classes.update(us_codes)
            unique_international_classes.update(intl_codes)

            # Count total (including duplicates)
            total_us_classes += len(us_codes)
            total_international_classes += len(intl_codes)

            all_data.append({
                'serial_number': sn,
                'mark_name': mark_name,
                'filing_date': class_data.get('filing_date', ''),
                'us_classes': class_data['us_classes'],
                'international_classes': class_data['international_classes'],
                'us_class_codes': ', '.join(us_codes),
                'international_class_codes': ', '.join(intl_codes),
                'description': class_data['description'],
                'mark_type': class_data.get('mark_type', 0),
                'error': class_data.get('error', None)
            })

            # Rate limiting: add delay between requests to avoid overwhelming the API
            time.sleep(0.75)  # Increased from 0.3s to 0.75s for better rate limiting

        # Log summary of errors if any occurred
        if failed_serials:
            print(f"\n⚠ WARNING: {len(failed_serials)} serial number(s) failed to load:")
            for failed in failed_serials:
                print(f"  ✗ {failed['serial_number']} ({failed['mark_name']}): {failed['error']}")
            print(f"✓ Successfully loaded: {len(serials) - len(failed_serials)}/{len(serials)}")
        else:
            print(f"\n✓ All {len(serials)} serial numbers loaded successfully!")

        return {
            'opposition_number': opposition_number,
            'serial_count': len(serials),
            'data': all_data,
            'unique_us_classes': sorted(unique_us_classes),
            'unique_international_classes': sorted(unique_international_classes),
            'total_us_classes': total_us_classes,
            'total_international_classes': total_international_classes,
            'filing_date': result_info.get('filing_date', ''),
            'termination_date': result_info.get('termination_date', ''),
            'result': result_info.get('result', None),
            'failed_serials': failed_serials  # Include error summary in return data
        }

    def search_oppositions_by_party(self, party_name: str, start_date: str = None, end_date: str = None) -> List[Dict[str, str]]:
        """
        Search for oppositions by party name and filter by opposition filing date range.
        Returns list of opposition numbers with their filing dates.
        Handles pagination to fetch all results across multiple pages.
        """
        oppositions = []
        page = 1
        max_pages = 20  # Limit to prevent infinite loops (20 pages * 25 results = 500 max)

        while page <= max_pages:
            params = {
                'qt': 'adv',
                'pn': party_name,
                'procstatus': 'All',
                'page': page
            }

            try:
                response = requests.get(self.ttabvue_base_url, params=params, timeout=30)
                response.raise_for_status()
            except requests.RequestException as e:
                st.error(f"Error fetching party search results (page {page}): {e}")
                break

            soup = BeautifulSoup(response.text, 'html.parser')
            page_oppositions = []

            # Find all links with opposition proceedings
            for link in soup.find_all('a', href=lambda x: x and 'pno=' in x):
                href = link.get('href', '')

                # Extract proceeding number and type
                pno_match = re.search(r'pno=(\d+)', href)
                pty_match = re.search(r'pty=([A-Z]+)', href)

                if pno_match and pty_match:
                    opp_number = pno_match.group(1)
                    proc_type = pty_match.group(1)

                    # ONLY process oppositions (pty=OPP), skip other types
                    if proc_type != 'OPP':
                        continue

                    # Extract filing date from parent TD
                    # Date is concatenated with proceeding number: "9123951102/14/2018"
                    parent_td = link.find_parent('td')
                    opposition_date = None

                    if parent_td:
                        td_text = parent_td.get_text(strip=True)
                        # Extract date pattern MM/DD/YYYY from the TD
                        date_match = re.search(r'(\d{2}/\d{2}/\d{4})', td_text)
                        if date_match:
                            opposition_date = date_match.group(1)

                    # Filter by opposition date range if specified
                    if opposition_date:
                        include = True
                        if start_date or end_date:
                            try:
                                opp_datetime = datetime.strptime(opposition_date, '%m/%d/%Y')

                                if start_date:
                                    start_datetime = datetime.strptime(start_date, '%m/%d/%Y')
                                    if opp_datetime < start_datetime:
                                        include = False

                                if end_date:
                                    end_datetime = datetime.strptime(end_date, '%m/%d/%Y')
                                    if opp_datetime > end_datetime:
                                        include = False
                            except ValueError:
                                # Skip if date parsing fails
                                include = False

                        if include:
                            page_oppositions.append({
                                'opposition_number': opp_number,
                                'opposition_date': opposition_date
                            })
                    else:
                        # If no date found, include it anyway (will be filtered out if dates required)
                        if not start_date and not end_date:
                            page_oppositions.append({
                                'opposition_number': opp_number,
                                'opposition_date': None
                            })

            # If no oppositions found on this page, we've reached the end
            if not page_oppositions:
                # Check if there's a "Next" link to see if there are more pages
                next_link = soup.find('a', string='Next')
                if not next_link:
                    break  # No more pages

            oppositions.extend(page_oppositions)

            # Check if this is the last page by looking for "Next" link
            next_link = soup.find('a', string='Next')
            if not next_link:
                break  # No "Next" link means this is the last page

            page += 1
            time.sleep(0.5)  # Be polite to the server

        # Remove duplicates
        seen = set()
        unique_oppositions = []
        for opp in oppositions:
            if opp['opposition_number'] not in seen:
                seen.add(opp['opposition_number'])
                unique_oppositions.append(opp)

        return unique_oppositions

    def search_oppositions_with_retry(self, party_name: str, start_date: str = None, end_date: str = None, max_retries: int = 3) -> List[Dict[str, str]]:
        """
        Search for oppositions with retry logic for network errors.
        Used by batch validation to handle connection issues gracefully.

        Args:
            party_name: Company name to search
            start_date: Optional start date filter (MM/DD/YYYY)
            end_date: Optional end date filter (MM/DD/YYYY)
            max_retries: Maximum number of retry attempts (default: 3)

        Returns:
            List of opposition numbers with dates, or empty list on failure
        """
        for attempt in range(max_retries):
            try:
                # Use the main search method
                oppositions = self.search_oppositions_by_party(party_name, start_date, end_date)
                return oppositions
            except requests.exceptions.ConnectionError as e:
                # Connection reset or similar network error
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2  # Exponential backoff: 2s, 4s, 6s
                    time.sleep(wait_time)
                    continue
                else:
                    # All retries exhausted
                    return []
            except requests.exceptions.Timeout as e:
                # Timeout error
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2
                    time.sleep(wait_time)
                    continue
                else:
                    return []
            except Exception as e:
                # Other unexpected errors
                return []

        return []

    def quick_check_oppositions(self, party_name: str, max_retries: int = 3) -> Dict:
        """
        Quick check to see if a company has ANY oppositions (optimized for batch validation).
        Only checks page 1, doesn't filter by dates - just detects if oppositions exist.

        This is much faster than full search because:
        - Only fetches page 1 (not all pages)
        - Doesn't filter by date range
        - Returns immediately when oppositions found

        Args:
            party_name: Company name to search
            max_retries: Maximum retry attempts for network errors

        Returns:
            Dict with 'has_oppositions' (bool), 'count' (int), 'sample_dates' (list)
        """
        for attempt in range(max_retries):
            try:
                params = {
                    'qt': 'adv',
                    'pn': party_name,
                    'procstatus': 'All',
                    'page': 1  # Only check first page
                }

                response = requests.get(self.ttabvue_base_url, params=params, timeout=30)
                response.raise_for_status()

                soup = BeautifulSoup(response.text, 'html.parser')
                opposition_dates = []

                # Find all links with opposition proceedings
                for link in soup.find_all('a', href=lambda x: x and 'pno=' in x):
                    href = link.get('href', '')
                    pty_match = re.search(r'pty=([A-Z]+)', href)

                    if pty_match:
                        proc_type = pty_match.group(1)

                        # Only count oppositions (OPP type)
                        if proc_type == 'OPP':
                            # Try to extract date
                            parent_td = link.find_parent('td')
                            if parent_td:
                                td_text = parent_td.get_text(strip=True)
                                date_match = re.search(r'(\d{2}/\d{2}/\d{4})', td_text)
                                if date_match:
                                    opposition_dates.append(date_match.group(1))

                # Return results
                has_opps = len(opposition_dates) > 0
                return {
                    'has_oppositions': has_opps,
                    'count': len(opposition_dates),
                    'sample_dates': opposition_dates[:3] if opposition_dates else []  # First 3 dates as samples
                }

            except requests.exceptions.ConnectionError as e:
                if attempt < max_retries - 1:
                    time.sleep((attempt + 1) * 2)  # Exponential backoff
                    continue
                else:
                    return {'has_oppositions': False, 'count': 0, 'sample_dates': [], 'error': 'Connection error'}

            except requests.exceptions.Timeout as e:
                if attempt < max_retries - 1:
                    time.sleep((attempt + 1) * 2)
                    continue
                else:
                    return {'has_oppositions': False, 'count': 0, 'sample_dates': [], 'error': 'Timeout'}

            except Exception as e:
                return {'has_oppositions': False, 'count': 0, 'sample_dates': [], 'error': str(e)}

        return {'has_oppositions': False, 'count': 0, 'sample_dates': [], 'error': 'Max retries exceeded'}

    def search_proceedings_from_url(self, url: str, start_date: str = None, end_date: str = None) -> List[Dict[str, str]]:
        """
        Extract proceeding numbers from a TTABVue search URL.
        Filter by proceeding filing date and type (oppositions only).
        Returns list of opposition numbers with their filing dates.
        """
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()
        except requests.RequestException as e:
            st.error(f"Error fetching URL: {e}")
            return []

        soup = BeautifulSoup(response.text, 'html.parser')
        proceedings = []

        # Find all links with pno= parameter
        for link in soup.find_all('a', href=lambda x: x and 'pno=' in x):
            href = link.get('href', '')

            # Extract proceeding number
            pno_match = re.search(r'pno=(\d+)', href)
            # Extract proceeding type (pty=)
            pty_match = re.search(r'pty=([A-Z]+)', href)

            if pno_match:
                proc_number = pno_match.group(1)
                proc_type = pty_match.group(1) if pty_match else None

                # Only process oppositions (OPP)
                if proc_type != 'OPP':
                    continue

                # Extract proceeding filing date
                # The date is in the same cell as the link, after a <br> tag
                # Get the parent td element
                parent_td = link.find_parent('td')
                filing_date = None

                if parent_td:
                    # Get all text from the td and look for date
                    # Date format: MM/DD/YYYY (no word boundary needed due to concatenation)
                    td_text = parent_td.get_text()
                    date_match = re.search(r'(\d{2}/\d{2}/\d{4})', td_text)
                    if date_match:
                        filing_date = date_match.group(1)

                # Filter by proceeding filing date range
                if filing_date:
                    include = True
                    if start_date or end_date:
                        try:
                            filing_datetime = datetime.strptime(filing_date, '%m/%d/%Y')

                            if start_date:
                                start_datetime = datetime.strptime(start_date, '%m/%d/%Y')
                                if filing_datetime < start_datetime:
                                    include = False

                            if end_date:
                                end_datetime = datetime.strptime(end_date, '%m/%d/%Y')
                                if filing_datetime > end_datetime:
                                    include = False
                        except ValueError:
                            include = False

                    if include:
                        proceedings.append({
                            'proceeding_number': proc_number,
                            'proceeding_type': proc_type,
                            'filing_date': filing_date
                        })
                else:
                    # Include if no date filtering
                    if not start_date and not end_date:
                        proceedings.append({
                            'proceeding_number': proc_number,
                            'proceeding_type': proc_type,
                            'filing_date': None
                        })

        # Remove duplicates
        seen = set()
        unique_proceedings = []
        for proc in proceedings:
            if proc['proceeding_number'] not in seen:
                seen.add(proc['proceeding_number'])
                unique_proceedings.append(proc)

        return unique_proceedings

    def scrape_oppositions_from_url(self, url: str, start_date: str = None, end_date: str = None, progress_callback=None) -> Dict:
        """
        Scrape all oppositions from a party URL within a proceeding filing date range.
        Only processes oppositions (OPP), excludes cancellations and other proceeding types.
        Returns aggregated results from all oppositions.
        """
        if progress_callback:
            progress_callback(0, "Extracting oppositions from URL...")

        # Get list of oppositions from URL
        proceedings = self.search_proceedings_from_url(url, start_date, end_date)

        if not proceedings:
            return {
                'url': url,
                'opposition_count': 0,
                'total_serial_count': 0,
                'data': [],
                'unique_us_classes': [],
                'unique_international_classes': []
            }

        # Aggregate data from all oppositions
        all_data = []
        global_unique_us_classes = set()
        global_unique_international_classes = set()
        global_total_us_classes = 0
        global_total_international_classes = 0

        total_proceedings = len(proceedings)

        for idx, proc_info in enumerate(proceedings, 1):
            proc_number = proc_info['proceeding_number']
            filing_date = proc_info.get('filing_date', '')

            if progress_callback:
                progress_callback(
                    idx / total_proceedings,
                    f"Processing opposition {idx}/{total_proceedings}: {proc_number}"
                )

            # Scrape this opposition
            result = self.scrape_opposition(proc_number, 'OPP')

            # Add proceeding info to each row
            for item in result['data']:
                item['proceeding_number'] = proc_number
                item['proceeding_filing_date'] = filing_date
                all_data.append(item)

            # Update global unique classes
            global_unique_us_classes.update(result['unique_us_classes'])
            global_unique_international_classes.update(result['unique_international_classes'])
            global_total_us_classes += result.get('total_us_classes', 0)
            global_total_international_classes += result.get('total_international_classes', 0)

            # Rate limiting
            time.sleep(0.5)

        return {
            'url': url,
            'opposition_count': total_proceedings,
            'total_serial_count': len(all_data),
            'data': all_data,
            'unique_us_classes': sorted(global_unique_us_classes),
            'unique_international_classes': sorted(global_unique_international_classes),
            'total_us_classes': global_total_us_classes,
            'total_international_classes': global_total_international_classes
        }

    def scrape_party_oppositions(self, party_name: str, start_date: str = None, end_date: str = None, progress_callback=None) -> Dict:
        """
        Scrape all oppositions for a party within a date range.
        Returns aggregated results from all oppositions.
        """
        if progress_callback:
            progress_callback(0, f"Searching oppositions for {party_name}...")

        # Get list of oppositions
        oppositions = self.search_oppositions_by_party(party_name, start_date, end_date)

        if not oppositions:
            return {
                'party_name': party_name,
                'opposition_count': 0,
                'total_serial_count': 0,
                'data': [],
                'unique_us_classes': [],
                'unique_international_classes': []
            }

        # Aggregate data from all oppositions
        all_data = []
        global_unique_us_classes = set()
        global_unique_international_classes = set()
        global_total_us_classes = 0
        global_total_international_classes = 0

        total_oppositions = len(oppositions)

        for idx, opp_info in enumerate(oppositions, 1):
            opp_number = opp_info['opposition_number']

            if progress_callback:
                progress_callback(
                    idx / total_oppositions,
                    f"Processing opposition {idx}/{total_oppositions}: {opp_number}"
                )

            # Scrape this opposition
            result = self.scrape_opposition(opp_number, 'OPP')

            # Add opposition number and opposition date to each row
            for item in result['data']:
                item['opposition_number'] = opp_number
                item['opposition_date'] = opp_info.get('opposition_date', '')
                all_data.append(item)

            # Update global unique classes
            global_unique_us_classes.update(result['unique_us_classes'])
            global_unique_international_classes.update(result['unique_international_classes'])
            global_total_us_classes += result.get('total_us_classes', 0)
            global_total_international_classes += result.get('total_international_classes', 0)

            # Rate limiting
            time.sleep(0.5)

        return {
            'party_name': party_name,
            'opposition_count': total_oppositions,
            'total_serial_count': len(all_data),
            'data': all_data,
            'unique_us_classes': sorted(global_unique_us_classes),
            'unique_international_classes': sorted(global_unique_international_classes),
            'total_us_classes': global_total_us_classes,
            'total_international_classes': global_total_international_classes
        }

    def bulk_scrape_party_consolidated(self, party_name: str, start_date: str = None, end_date: str = None, progress_callback=None) -> Dict:
        """
        Bulk scrape all oppositions for a party with ONE ROW per opposition.

        Each opposition's data is consolidated into a single row with:
        - All serial numbers as comma-separated list
        - All mark names as comma-separated list
        - Unique and total class counts
        - Error tracking for failed serials

        Args:
            party_name: Name of the party to search
            start_date: Optional start date filter (MM/DD/YYYY)
            end_date: Optional end date filter (MM/DD/YYYY)
            progress_callback: Callback function for progress updates

        Returns:
            Dict with:
                - party_name: str
                - search_date: str (timestamp)
                - opposition_count: int
                - oppositions: List[Dict] - one dict per opposition
                - processing_stats: Dict with timing and cache statistics
        """
        import time as time_module
        from datetime import datetime

        start_time = time_module.time()

        if progress_callback:
            progress_callback(0, f"Searching oppositions for {party_name}...")

        # STEP 1: Search for all opposition numbers
        oppositions = self.search_oppositions_by_party(party_name, start_date, end_date)

        if not oppositions:
            return {
                'party_name': party_name,
                'search_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'opposition_count': 0,
                'oppositions': [],
                'processing_stats': {
                    'total_time_seconds': 0,
                    'cache_hits': 0,
                    'cache_misses': 0,
                    'errors': 0
                }
            }

        # STEP 2: Process each opposition sequentially
        opposition_rows = []
        total_oppositions = len(oppositions)
        total_errors = 0
        initial_cache_hits = self.session_stats['cache_hits']
        initial_cache_misses = self.session_stats['cache_misses']

        for idx, opp_info in enumerate(oppositions, 1):
            opp_number = opp_info['opposition_number']
            opp_date = opp_info.get('opposition_date', '')

            if progress_callback:
                progress_callback(
                    idx / total_oppositions,
                    f"Processing opposition {idx}/{total_oppositions}: {opp_number}"
                )

            try:
                # STEP 3: Scrape this opposition using existing method
                result = self.scrape_opposition(opp_number, 'OPP')

                # STEP 4: Consolidate data into ONE ROW
                row = self._consolidate_opposition_to_row(
                    opposition_number=opp_number,
                    opposition_date=opp_date,
                    result=result
                )

                opposition_rows.append(row)

                # Rate limiting between oppositions
                if idx < total_oppositions:
                    time.sleep(0.75)

            except Exception as e:
                # Log error but continue with next opposition
                print(f"✗ Error processing opposition {opp_number}: {str(e)}")
                total_errors += 1

                # Add error row
                opposition_rows.append({
                    'opposition_number': opp_number,
                    'opposition_date': opp_date,
                    'status': 'FAILED',
                    'error_message': str(e),
                    'serial_count': 0,
                    'serial_numbers': '',
                    'mark_names': '',
                    'filing_dates': '',
                    'all_us_classes': '',
                    'unique_us_classes': '',
                    'total_us_class_count': 0,
                    'all_international_classes': '',
                    'unique_international_classes': '',
                    'total_international_class_count': 0,
                    'mark_types': '',
                    'failed_serials': '',
                    'cache_hit_rate': '0%'
                })

        end_time = time_module.time()
        total_time = end_time - start_time

        # Calculate cache statistics
        cache_hits_session = self.session_stats['cache_hits'] - initial_cache_hits
        cache_misses_session = self.session_stats['cache_misses'] - initial_cache_misses

        return {
            'party_name': party_name,
            'search_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'opposition_count': len(opposition_rows),
            'oppositions': opposition_rows,
            'processing_stats': {
                'total_time_seconds': round(total_time, 2),
                'time_per_opposition': round(total_time / len(opposition_rows), 2) if opposition_rows else 0,
                'cache_hits': cache_hits_session,
                'cache_misses': cache_misses_session,
                'cache_hit_rate': round(cache_hits_session / (cache_hits_session + cache_misses_session) * 100, 1) if (cache_hits_session + cache_misses_session) > 0 else 0,
                'total_errors': total_errors,
                'successful_oppositions': len([r for r in opposition_rows if r.get('status') != 'FAILED'])
            }
        }

    def _consolidate_opposition_to_row(self, opposition_number: str, opposition_date: str, result: Dict) -> Dict:
        """
        Helper method to consolidate opposition data into a single row.

        Takes the result from scrape_opposition() and creates a single row with:
        - Comma-separated serial numbers
        - Comma-separated mark names
        - Comma-separated filing dates
        - All classes and unique classes
        - Mark types
        - Error tracking
        """
        data = result.get('data', [])

        # Extract lists
        serial_numbers = [item['serial_number'] for item in data]
        mark_names = [item['mark_name'] for item in data]
        filing_dates = [item.get('filing_date', '') for item in data]
        mark_types = [str(item.get('mark_type', 0)) for item in data]

        # Collect all classes (with duplicates for counting)
        all_us_classes = []
        all_intl_classes = []

        for item in data:
            us_codes = [c['code'] for c in item.get('us_classes', [])]
            intl_codes = [c['code'] for c in item.get('international_classes', [])]
            all_us_classes.extend(us_codes)
            all_intl_classes.extend(intl_codes)

        # Get unique classes (sorted)
        unique_us = sorted(set(all_us_classes))
        unique_intl = sorted(set(all_intl_classes))

        # Track failed serials
        failed_serials = [
            f"{item['serial_number']} ({item.get('error', 'Unknown error')})"
            for item in data
            if item.get('error')
        ]

        # Calculate cache hit rate for this opposition
        total_serials = len(serial_numbers)
        cached_serials = len([item for item in data if item.get('from_cache', False)])
        cache_hit_rate = f"{(cached_serials / total_serials * 100):.1f}%" if total_serials > 0 else "0%"

        # Get party information
        party_info = self.get_party_info(opposition_number)

        return {
            'opposition_number': opposition_number,
            'filing_date': result.get('filing_date', ''),
            'termination_date': result.get('termination_date', ''),
            'status': 'SUCCESS',
            'result': result.get('result', None),  # 0=Dismissed, 1=Sustained, None=Pending
            'plaintiff': party_info.get('plaintiff_name', ''),
            'defendant': party_info.get('defendant_name', ''),
            'serial_count': len(serial_numbers),
            'serial_numbers': ', '.join(serial_numbers),
            'mark_names': ', '.join(mark_names),
            'filing_dates': ', '.join(filing_dates),
            'all_us_classes': ', '.join(all_us_classes),
            'unique_us_classes': ', '.join(unique_us),
            'total_us_class_count': len(all_us_classes),
            'all_international_classes': ', '.join(all_intl_classes),
            'unique_international_classes': ', '.join(unique_intl),
            'total_international_class_count': len(all_intl_classes),
            'mark_types': ', '.join(mark_types),
            'failed_serials': '; '.join(failed_serials) if failed_serials else '',
            'cache_hit_rate': cache_hit_rate
        }

    def batch_analyze_oppositions(self, url: str, company_name: str, gvkey: str = None,
                                  start_date: str = None, end_date: str = None,
                                  progress_callback=None) -> Dict:
        """
        Comprehensive batch analysis of oppositions from URL.
        Returns data in the format matching image1.png Excel structure.
        One row per opposition with all required fields.
        """
        if progress_callback:
            progress_callback(0, "Extracting oppositions from URL...")

        # Get list of oppositions
        proceedings = self.search_proceedings_from_url(url, start_date, end_date)

        if not proceedings:
            return {
                'company_name': company_name,
                'gvkey': gvkey,
                'opposition_count': 0,
                'data': []
            }

        total_proceedings = len(proceedings)
        all_opposition_data = []

        for idx, proc_info in enumerate(proceedings, 1):
            proc_number = proc_info['proceeding_number']

            if progress_callback:
                progress_callback(
                    idx / total_proceedings,
                    f"Analyzing opposition {idx}/{total_proceedings}: {proc_number}"
                )

            # Analyze this opposition comprehensively
            opp_analysis = self.analyze_opposition_complete(
                proc_number,
                company_name,
                'OPP'
            )

            # Add GVKEY
            opp_analysis['gvkey'] = gvkey
            opp_analysis['company_name'] = company_name

            all_opposition_data.append(opp_analysis)

            # Rate limiting
            time.sleep(0.5)

        return {
            'company_name': company_name,
            'gvkey': gvkey,
            'opposition_count': total_proceedings,
            'data': all_opposition_data
        }


def create_excel_file(result: Dict, is_party_search: bool = False) -> bytes:
    """Create Excel file in memory and return bytes."""
    output = io.BytesIO()

    rows = []
    for item in result['data']:
        mark_type_label = {
            0: 'No Image',
            1: 'Standard Text',
            2: 'Stylized/Design',
            3: 'Slogan'
        }.get(item.get('mark_type', 0), 'No Image')

        row_data = {
            'Serial Number': item['serial_number'],
            'Mark Name': item['mark_name'],
            'Filing Date': item.get('filing_date', ''),
            'Mark Type': item.get('mark_type', 0),
            'Mark Type Label': mark_type_label,
            'US Classes': item['us_class_codes'],
            'International Classes': item['international_class_codes'],
            'Description': item['description']
        }

        # Add opposition/proceeding number column if party search
        if is_party_search:
            # Check if we have proceeding_number (URL-based search) or opposition_number (party name search)
            proc_num = item.get('proceeding_number', item.get('opposition_number', ''))
            proc_date = item.get('proceeding_filing_date', item.get('opposition_date', ''))
            row_data = {'Proceeding Number': proc_num, 'Proceeding Filing Date': proc_date, **row_data}
        else:
            row_data = {'Opposition Number': result.get('opposition_number', ''), **row_data}

        rows.append(row_data)

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Trademark Classes', index=False)

        # Create summary based on search type
        if is_party_search:
            summary_data = {
                'Metric': [
                    'Party Name',
                    'Total Oppositions',
                    'Total Serial Numbers',
                    'Unique US Classes',
                    'Total US Classes Count',
                    'Unique International Classes',
                    'Total International Classes Count'
                ],
                'Value': [
                    result.get('party_name', ''),
                    result.get('opposition_count', 0),
                    result.get('total_serial_count', 0),
                    ', '.join(result['unique_us_classes']),
                    result.get('total_us_classes', 0),
                    ', '.join(result['unique_international_classes']),
                    result.get('total_international_classes', 0)
                ]
            }
        else:
            summary_data = {
                'Metric': [
                    'Opposition Number',
                    'Total Serial Numbers',
                    'Unique US Classes',
                    'Total US Classes Count',
                    'Unique International Classes',
                    'Total International Classes Count'
                ],
                'Value': [
                    result.get('opposition_number', ''),
                    result.get('serial_count', 0),
                    ', '.join(result['unique_us_classes']),
                    result.get('total_us_classes', 0),
                    ', '.join(result['unique_international_classes']),
                    result.get('total_international_classes', 0)
                ]
            }

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    return output.getvalue()


def create_comprehensive_excel(result: Dict) -> bytes:
    """
    Create Excel file matching image1.png format.
    One row per opposition with separate columns for each serial number.
    """
    output = io.BytesIO()

    rows = []
    max_marks = 0  # Track maximum number of marks across all oppositions

    # First pass: determine max number of marks
    for opp in result['data']:
        max_marks = max(max_marks, len(opp.get('mark_details', [])))

    # Second pass: build rows
    for opp in result['data']:
        row_data = {
            'GVKEY': opp.get('gvkey', ''),
            'C': opp.get('company_name', ''),
            'Alt Name': opp.get('alt_name', ''),
            'Plaintiff': opp.get('plaintiff', 0),
            'Marks': opp.get('marks', 0),
            'US GS': opp.get('us_gs', 0),
            'INT GS': opp.get('int_gs', 0),
            'Opp Start Date': opp.get('opp_start_date', ''),
            'Opp End Date': opp.get('opp_end_date', ''),
            'Result': opp.get('result', ''),
            'TM Type': opp.get('tm_type_1', 0)  # Could expand to show all types
        }

        # Add serial number and trademark columns
        mark_details = opp.get('mark_details', [])
        for idx, mark in enumerate(mark_details):
            col_num = idx + 1
            row_data[f'Serial No {col_num}'] = mark['serial_number']
            row_data[f'Trademark {col_num}'] = mark['mark_name']

        # Fill remaining columns with empty values
        for idx in range(len(mark_details), max_marks):
            col_num = idx + 1
            row_data[f'Serial No {col_num}'] = ''
            row_data[f'Trademark {col_num}'] = ''

        rows.append(row_data)

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Opposition Analysis', index=False)

    return output.getvalue()


def create_bulk_party_excel(result: Dict) -> bytes:
    """
    Create Excel file for bulk party scraping with TWO SHEETS:
    - Sheet 1: "Opposition Data" - One row per opposition (visible)
    - Sheet 2: "Metadata" - Processing statistics (hidden)

    Each row contains ALL data for one opposition consolidated:
    - Opposition info, plaintiff, defendant
    - Comma-separated serial numbers and mark names
    - All classes and unique classes
    - Mark types and error tracking
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    output = io.BytesIO()

    # SHEET 1: Opposition Data (one row per opposition)
    opposition_rows = []

    for opp in result.get('oppositions', []):
        # Map result code to text
        result_code = opp.get('result')
        if result_code == 1:
            result_text = 'Sustained'
        elif result_code == 0:
            result_text = 'Dismissed'
        else:
            result_text = 'Pending'

        row = {
            'Opposition Number': opp.get('opposition_number', ''),
            'Filing Date': opp.get('filing_date', ''),
            'Termination Date': opp.get('termination_date', ''),
            'Status': opp.get('status', ''),
            'Result': result_text,
            'Plaintiff': opp.get('plaintiff', ''),
            'Defendant': opp.get('defendant', ''),
            'Serial Count': opp.get('serial_count', 0),
            'Serial Numbers': opp.get('serial_numbers', ''),
            'Mark Names': opp.get('mark_names', ''),
            'Filing Dates': opp.get('filing_dates', ''),
            'All US Classes': opp.get('all_us_classes', ''),
            'Unique US Classes': opp.get('unique_us_classes', ''),
            'Total US Class Count': opp.get('total_us_class_count', 0),
            'All International Classes': opp.get('all_international_classes', ''),
            'Unique International Classes': opp.get('unique_international_classes', ''),
            'Total International Class Count': opp.get('total_international_class_count', 0),
            'Mark Types': opp.get('mark_types', ''),
            'Cache Hit Rate': opp.get('cache_hit_rate', '0%'),
            'Failed Serials': opp.get('failed_serials', ''),
            'Error Message': opp.get('error_message', '')
        }
        opposition_rows.append(row)

    df_data = pd.DataFrame(opposition_rows)

    # SHEET 2: Metadata (hidden)
    stats = result.get('processing_stats', {})
    metadata = {
        'Metric': [
            'Party Name',
            'Search Date',
            'Total Oppositions Scraped',
            'Successful Oppositions',
            'Failed Oppositions',
            'Total Processing Time (seconds)',
            'Average Time per Opposition (seconds)',
            'Cache Hits',
            'Cache Misses',
            'Cache Hit Rate (%)',
            'API Calls Saved',
            'Estimated Cost Saved (Anthropic)',
        ],
        'Value': [
            result.get('party_name', ''),
            result.get('search_date', ''),
            result.get('opposition_count', 0),
            stats.get('successful_oppositions', 0),
            stats.get('total_errors', 0),
            stats.get('total_time_seconds', 0),
            stats.get('time_per_opposition', 0),
            stats.get('cache_hits', 0),
            stats.get('cache_misses', 0),
            stats.get('cache_hit_rate', 0),
            stats.get('cache_hits', 0),
            f"${stats.get('cache_hits', 0) * 0.015:.2f}"
        ]
    }
    df_metadata = pd.DataFrame(metadata)

    # Write to Excel with openpyxl for advanced formatting
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write Sheet 1: Opposition Data
        df_data.to_excel(writer, sheet_name='Opposition Data', index=False)

        # Write Sheet 2: Metadata
        df_metadata.to_excel(writer, sheet_name='Metadata', index=False)

        # Access workbook and sheets
        workbook = writer.book
        data_sheet = workbook['Opposition Data']
        metadata_sheet = workbook['Metadata']

        # Format Sheet 1: Opposition Data
        # Header styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        for cell in data_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Auto-adjust column widths
        for column in data_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long columns
            data_sheet.column_dimensions[column_letter].width = adjusted_width

        # Format Sheet 2: Metadata
        for cell in metadata_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Bold the metric names
        for row in metadata_sheet.iter_rows(min_row=2, max_row=metadata_sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)

        # Auto-adjust column widths for metadata
        for column in metadata_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            metadata_sheet.column_dimensions[column_letter].width = adjusted_width

        # Hide the Metadata sheet
        metadata_sheet.sheet_state = 'hidden'

    return output.getvalue()


def main():
    """Main Streamlit app."""

    st.set_page_config(
        page_title="USPTO Opposition Scraper",
        page_icon="⚖️",
        layout="wide"
    )

    st.title("⚖️ USPTO Opposition Trademark Class Scraper")
    st.markdown("Retrieve US and International classes from opposition pleaded applications")

    # Sidebar - Cache Statistics
    with st.sidebar:
        st.header("📊 Cache Statistics")

        # Get cache statistics
        cache = TrademarkCache()
        stats = cache.get_cache_statistics()

        st.metric("Total Cached Records", stats['total_cached_records'])
        st.metric("Cache Hit Rate (24h)", f"{stats['hit_rate_24h']:.1f}%")
        st.metric("Anthropic API Calls Saved", stats['anthropic_calls_saved'])
        st.metric("TSDR API Calls Saved", stats['tsdr_calls_saved'])

        # Calculate estimated savings
        estimated_savings = stats['anthropic_calls_saved'] * 0.015
        st.metric("Estimated Cost Savings", f"${estimated_savings:.2f}")

        st.divider()

        # Cache management buttons
        st.subheader("🔧 Cache Management")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Clear Stale", help="Clear records older than 30 days"):
                deleted = cache.clear_stale_records()
                st.success(f"Cleared {deleted} stale records")

        with col2:
            if st.button("Clear All", help="Clear entire cache"):
                cache.clear_all_cache()
                st.success("Cache cleared!")

    # Configuration - Load API keys from environment variables
    API_KEY = os.getenv("USPTO_API_KEY", "")
    ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
    CLAUDE_VISION_API_KEY = ANTHROPIC_API_KEY  # Same key used for vision

    # Check if API keys are set
    if not API_KEY:
        st.error("❌ USPTO_API_KEY not found in environment variables. Please set up your .env file.")
        st.stop()

    if not ANTHROPIC_API_KEY:
        st.error("❌ ANTHROPIC_API_KEY not found in environment variables. Please set up your .env file.")
        st.stop()

    # Create tabs for different search modes
    tab1, tab2, tab3 = st.tabs(["🔍 Single Opposition Search", "📦 Bulk Party Search", "✅ Batch Company Validation"])

    # TAB 1: Single Opposition Search
    with tab1:
        st.markdown("### Search by Opposition Number")
        st.markdown("Enter a single opposition number to retrieve trademark class data.")

        # Input section
        col1, col2 = st.columns([3, 1])

    with col1:
        opposition_number = st.text_input(
            "Enter Opposition Number",
            placeholder="e.g., 91302017",
            help="Enter the USPTO opposition number to retrieve trademark class data"
        )

    with col2:
        st.write("")
        st.write("")
        search_button = st.button("🔍 Search", type="primary", use_container_width=True)

    # Process when search button is clicked
    if search_button and opposition_number:

        scraper = USPTOOppositionScraper(API_KEY, CLAUDE_VISION_API_KEY, CLAUDE_VISION_API_KEY)

        # Progress tracking
        progress_bar = st.progress(0)
        status_text = st.empty()

        def update_progress(progress, message):
            progress_bar.progress(progress)
            status_text.text(message)

        # Scrape data
        with st.spinner("Fetching opposition data..."):
            result = scraper.scrape_opposition(opposition_number, 'OPP', update_progress)

        progress_bar.empty()
        status_text.empty()

        if result['serial_count'] == 0:
            st.error("❌ No serial numbers found in pleaded applications section. Please check the opposition number.")
        else:
            # Summary section
            st.success(f"✅ Found {result['serial_count']} serial numbers")

            # Metrics - First row
            col1, col2, col3 = st.columns(3)

            with col1:
                st.metric("Opposition Number", result['opposition_number'])

            with col2:
                st.metric("Total Serial Numbers", result['serial_count'])

            with col3:
                # Display result as text
                opp_result = result.get('result', None)
                if opp_result == 1:
                    result_text = "Sustained"
                    st.metric("Result", result_text, delta="Opposition Sustained", delta_color="normal")
                elif opp_result == 0:
                    result_text = "Dismissed"
                    st.metric("Result", result_text, delta="Opposition Dismissed", delta_color="inverse")
                else:
                    st.metric("Result", "Pending")

            # Metrics - Second row
            col1, col2 = st.columns(2)

            with col1:
                filing_date = result.get('filing_date', 'N/A')
                st.metric("Filing Date", filing_date if filing_date else 'N/A')

            with col2:
                termination_date = result.get('termination_date', 'N/A')
                st.metric("Termination Date", termination_date if termination_date else 'N/A')

            # Display unique classes
            col1, col2 = st.columns(2)
            with col1:
                us_classes_str = ', '.join(result['unique_us_classes']) if result['unique_us_classes'] else 'None'
                us_count = len(result['unique_us_classes'])
                st.info(f"**Unique US Classes ({us_count}):** {us_classes_str}")

            with col2:
                intl_classes_str = ', '.join(result['unique_international_classes']) if result['unique_international_classes'] else 'None'
                intl_count = len(result['unique_international_classes'])
                st.info(f"**Unique International Classes ({intl_count}):** {intl_classes_str}")

            # Display counts
            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Total US Classes Count", result.get('total_us_classes', 0))
            with col2:
                st.metric("Total International Classes Count", result.get('total_international_classes', 0))

            # Main data table
            st.subheader("📊 Trademark Classes Data")

            # Prepare DataFrame for display
            display_data = []
            for item in result['data']:
                mark_type_label = {
                    0: 'No Image',
                    1: '1 - Standard',
                    2: '2 - Stylized',
                    3: '3 - Slogan'
                }.get(item.get('mark_type', 0), 'No Image')

                display_data.append({
                    'Serial Number': item['serial_number'],
                    'Mark Name': item['mark_name'],
                    'Filing Date': item.get('filing_date', ''),
                    'Mark Type': mark_type_label,
                    'US Classes': item['us_class_codes'],
                    'International Classes': item['international_class_codes'],
                    'Description': item['description']
                })

            df = pd.DataFrame(display_data)

            # Display table
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Serial Number": st.column_config.TextColumn("Serial Number", width="medium"),
                    "Mark Name": st.column_config.TextColumn("Mark Name", width="medium"),
                    "Filing Date": st.column_config.TextColumn("Filing Date", width="small"),
                    "Mark Type": st.column_config.TextColumn("Mark Type", width="small"),
                    "US Classes": st.column_config.TextColumn("US Classes", width="small"),
                    "International Classes": st.column_config.TextColumn("International Classes", width="small"),
                    "Description": st.column_config.TextColumn("Description", width="large")
                }
            )

            # Download section
            st.subheader("💾 Download Results")

            col1, col2 = st.columns(2)

            with col1:
                # Excel download
                excel_data = create_excel_file(result)
                st.download_button(
                    label="📥 Download Excel",
                    data=excel_data,
                    file_name=f"opposition_{opposition_number}_classes.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with col2:
                # JSON download
                json_data = {
                    'opposition_number': result['opposition_number'],
                    'serial_count': result['serial_count'],
                    'unique_us_classes': result['unique_us_classes'],
                    'unique_international_classes': result['unique_international_classes'],
                    'trademarks': result['data']
                }
                json_str = json.dumps(json_data, indent=2)

                st.download_button(
                    label="📥 Download JSON",
                    data=json_str,
                    file_name=f"opposition_{opposition_number}_classes.json",
                    mime="application/json",
                    use_container_width=True
                )

            # Copyable text output section
            st.subheader("📋 Copyable Summary")

            # Create single row data
            row_data = []

            # Marks = total serial count
            marks = str(result['serial_count'])
            row_data.append(marks)

            # US GS = total US classes count
            us_gs = str(result.get('total_us_classes', 0))
            row_data.append(us_gs)

            # INT GS = total international classes count
            int_gs = str(result.get('total_international_classes', 0))
            row_data.append(int_gs)

            # Opp Start Date = filing date
            opp_start_date = result.get('filing_date', '') or ''
            row_data.append(str(opp_start_date))

            # Opp End Date = termination date
            opp_end_date = result.get('termination_date', '') or ''
            row_data.append(str(opp_end_date))

            # Result: 1=Sustained, 0=Dismissed, NA=Pending
            opp_result = result.get('result', None)
            result_text = "1" if opp_result == 1 else "0" if opp_result == 0 else "NA"
            row_data.append(result_text)

            # For each trademark, add mark_type and serial_number pairs
            for item in result['data']:
                tm_type = str(item.get('mark_type', 0))
                serial_no = str(item['serial_number'])
                row_data.append(tm_type)
                row_data.append(serial_no)

            # Create tab-separated text for clipboard (Excel format) - only data, no headers
            clipboard_text = "\t".join(row_data)

            # Display copy instructions
            st.write("**Copy the data below and paste it into Excel:**")
            st.info("💡 **How to paste in Excel:** Click on a single cell in Excel, then paste (Ctrl+V or Cmd+V). The data will automatically fill across the row.")

            # Show in a code block with copy button
            st.code(clipboard_text, language=None)

            # Streamlit's native copy to clipboard using pyperclip-style approach
            import streamlit.components.v1 as components

            # Create a button that copies to clipboard using HTML/JavaScript
            components.html(
                f"""
                <div style="margin: 10px 0;">
                    <button onclick="copyToClipboard()" style="
                        background-color: #FF4B4B;
                        color: white;
                        padding: 0.5rem 1rem;
                        border: none;
                        border-radius: 0.5rem;
                        cursor: pointer;
                        font-size: 1rem;
                        width: 100%;
                    ">
                        📋 Copy to Clipboard
                    </button>
                    <p id="status" style="margin-top: 10px; color: green;"></p>
                </div>
                <script>
                    function copyToClipboard() {{
                        const text = `{clipboard_text}`;
                        navigator.clipboard.writeText(text).then(function() {{
                            document.getElementById('status').textContent = '✅ Copied! Now paste into Excel (click one cell first)';
                            setTimeout(() => {{
                                document.getElementById('status').textContent = '';
                            }}, 3000);
                        }}, function(err) {{
                            document.getElementById('status').textContent = '❌ Copy failed. Please select and copy manually.';
                        }});
                    }}
                </script>
                """,
                height=100,
            )

            # Detailed view (expandable)
            with st.expander("🔍 View Detailed Class Information"):
                for item in result['data']:
                    st.markdown(f"### Serial Number: {item['serial_number']} - {item['mark_name']}")

                    if item['us_classes']:
                        st.markdown("**US Classes:**")
                        for uc in item['us_classes']:
                            st.markdown(f"- `{uc['code']}`: {uc['description']}")

                    if item['international_classes']:
                        st.markdown("**International Classes:**")
                        for ic in item['international_classes']:
                            st.markdown(f"- `{ic['code']}`: {ic['description']}")

                    if item['description']:
                        st.markdown(f"**Description:** {item['description']}")

                    st.divider()

    # TAB 2: Bulk Party Search
    with tab2:
        st.markdown("### Bulk Search by Party Name")
        st.markdown("Search for all oppositions involving a specific party and scrape them all at once.")
        st.info("⏱️ **Note:** This is a two-step process. First, we'll search for oppositions. Then you can confirm and start scraping.")
        st.info("📅 **Date Filtering:** The dates below filter by **Proceeding Filing Date** (when the opposition was filed with USPTO), not the opposition start/end dates within the proceeding.")

        # Initialize session state for storing search results
        if 'search_results' not in st.session_state:
            st.session_state.search_results = None
        if 'party_name_searched' not in st.session_state:
            st.session_state.party_name_searched = None

        # Input section
        col1, col2, col3 = st.columns([2, 1, 1])

        with col1:
            party_name = st.text_input(
                "Enter Party Name",
                placeholder="e.g., Nike Inc",
                help="Enter the company or party name to search for oppositions",
                key="party_name_input"
            )

        with col2:
            start_date = st.text_input(
                "Earliest Filing Date",
                value="01/01/2012",
                placeholder="MM/DD/YYYY",
                help="Only include oppositions filed AFTER this date (Proceeding Filing Date)",
                key="start_date_input"
            )

        with col3:
            end_date = st.text_input(
                "Latest Filing Date",
                value="12/31/2022",
                placeholder="MM/DD/YYYY",
                help="Only include oppositions filed BEFORE this date (Proceeding Filing Date)",
                key="end_date_input"
            )

        # Step 1: Search button
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            search_button = st.button("🔍 Step 1: Find Oppositions", type="primary", use_container_width=True)

        # Process search
        if search_button and party_name:
            # Validate date formats if provided
            date_valid = True
            if start_date:
                try:
                    datetime.strptime(start_date, '%m/%d/%Y')
                except ValueError:
                    st.error("❌ Invalid start date format. Please use MM/DD/YYYY")
                    date_valid = False

            if end_date:
                try:
                    datetime.strptime(end_date, '%m/%d/%Y')
                except ValueError:
                    st.error("❌ Invalid end date format. Please use MM/DD/YYYY")
                    date_valid = False

            if date_valid:
                # Create scraper just for searching
                scraper = USPTOOppositionScraper(API_KEY, CLAUDE_VISION_API_KEY, ANTHROPIC_API_KEY)

                with st.spinner(f"🔍 Searching for oppositions involving '{party_name}'..."):
                    try:
                        start_date_param = start_date if start_date else None
                        end_date_param = end_date if end_date else None

                        # Search for oppositions
                        oppositions = scraper.search_oppositions_by_party(party_name, start_date_param, end_date_param)

                        # Store results in session state
                        st.session_state.search_results = oppositions
                        st.session_state.party_name_searched = party_name

                    except Exception as e:
                        st.error(f"❌ Error searching for party: {str(e)}")
                        st.session_state.search_results = None
                        st.session_state.party_name_searched = None

        elif search_button and not party_name:
            st.error("❌ Please enter a party name to search")

        # Display search results
        if st.session_state.search_results is not None:
            oppositions = st.session_state.search_results

            if len(oppositions) == 0:
                st.warning(f"⚠️ No oppositions found for party: **{st.session_state.party_name_searched}**")
                st.info("💡 **Suggestions:**\n- Check the spelling of the party name\n- Try searching without date filters\n- Verify the party has oppositions in [TTABVue](https://ttabvue.uspto.gov)")

                # Clear session state
                if st.button("🔄 Try Another Search"):
                    st.session_state.search_results = None
                    st.session_state.party_name_searched = None
                    st.rerun()

            else:
                # Show success message
                st.success(f"✅ Found **{len(oppositions)}** opposition(s) for party: **{st.session_state.party_name_searched}**")

                # Display opposition list in a table
                st.markdown("---")
                st.subheader("📋 Oppositions Found")

                preview_data = []
                for idx, opp in enumerate(oppositions[:20], 1):  # Show first 20
                    preview_data.append({
                        '#': idx,
                        'Opposition Number': opp.get('opposition_number', ''),
                        'Filing Date': opp.get('opposition_date', 'N/A')
                    })

                df_preview = pd.DataFrame(preview_data)
                st.dataframe(df_preview, use_container_width=True, hide_index=True)

                if len(oppositions) > 20:
                    st.info(f"📝 Showing first 20 of {len(oppositions)} oppositions. All will be scraped if you proceed.")

                # Estimate time
                estimated_time_min = len(oppositions) * 0.5  # ~30 seconds per opposition on average
                estimated_time_max = len(oppositions) * 1.5  # ~90 seconds per opposition with no cache

                st.markdown("---")
                st.markdown("### ⏱️ Estimated Processing Time")
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("With Cache (Repeat Run)", f"{estimated_time_min:.0f}-{estimated_time_min*1.2:.0f} min")
                with col2:
                    st.metric("Without Cache (First Run)", f"{estimated_time_max:.0f}-{estimated_time_max*1.3:.0f} min")

                st.info("💡 **Tip:** Subsequent runs are much faster due to caching. Off-peak hours (10pm-5am EST) are ~2x faster.")

                # Step 2: Confirm and scrape
                st.markdown("---")
                st.markdown("### 🚀 Step 2: Confirm and Scrape")

                col1, col2, col3 = st.columns([1, 1, 2])
                with col1:
                    scrape_button = st.button("✅ Confirm & Start Scraping", type="primary", use_container_width=True, key="scrape_confirm")
                with col2:
                    if st.button("🔄 Cancel & New Search", use_container_width=True):
                        st.session_state.search_results = None
                        st.session_state.party_name_searched = None
                        st.rerun()

                # Process scraping
                if scrape_button:
                    scraper = USPTOOppositionScraper(API_KEY, CLAUDE_VISION_API_KEY, ANTHROPIC_API_KEY)

                    # Progress tracking
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    def update_progress(progress, message):
                        progress_bar.progress(progress)
                        status_text.text(message)

                    # Scrape data
                    try:
                        result = scraper.bulk_scrape_party_consolidated(
                            party_name=st.session_state.party_name_searched,
                            start_date=start_date if start_date else None,
                            end_date=end_date if end_date else None,
                            progress_callback=update_progress
                        )

                        progress_bar.empty()
                        status_text.empty()

                        if result['opposition_count'] == 0:
                            st.warning(f"❌ No data could be scraped")
                        else:
                            # Summary section
                            stats = result.get('processing_stats', {})
                            st.success(f"✅ Successfully scraped {stats.get('successful_oppositions', 0)} out of {result['opposition_count']} oppositions")

                        # Display metrics
                        col1, col2, col3, col4 = st.columns(4)
    
                        with col1:
                            st.metric("Total Oppositions", result['opposition_count'])
    
                        with col2:
                            st.metric("Processing Time", f"{stats.get('total_time_seconds', 0):.1f}s")
    
                        with col3:
                            st.metric("Cache Hit Rate", f"{stats.get('cache_hit_rate', 0):.1f}%")
    
                        with col4:
                            st.metric("Errors", stats.get('total_errors', 0))
    
                        # Show additional stats
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Avg Time per Opposition", f"{stats.get('time_per_opposition', 0):.1f}s")
                        with col2:
                            api_savings = stats.get('cache_hits', 0) * 0.015
                            st.metric("Est. Cost Saved", f"${api_savings:.2f}")
    
                        # Show preview of data
                        st.markdown("---")
                        st.subheader("📊 Opposition Data Preview")
    
                        # Create preview DataFrame
                        preview_data = []
                        for opp in result.get('oppositions', [])[:10]:  # Show first 10
                            preview_data.append({
                                'Opposition #': opp.get('opposition_number', ''),
                                'Status': opp.get('status', ''),
                                'Result': 'Sustained' if opp.get('result') == 1 else 'Dismissed' if opp.get('result') == 0 else 'Pending',
                                'Serial Count': opp.get('serial_count', 0),
                                'Unique US Classes': opp.get('unique_us_classes', ''),
                                'Unique Intl Classes': opp.get('unique_international_classes', ''),
                                'Failed Serials': opp.get('failed_serials', '')
                            })
    
                        if preview_data:
                            df_preview = pd.DataFrame(preview_data)
                            st.dataframe(df_preview, use_container_width=True, hide_index=True)
    
                            if result['opposition_count'] > 10:
                                st.info(f"📝 Showing first 10 of {result['opposition_count']} oppositions. Download the Excel file for complete data.")
    
                        # Download section
                        st.markdown("---")
                        st.subheader("💾 Download Complete Results")

                        # Export format selection
                        st.markdown("### Choose Export Format")
                        export_format = st.radio(
                            "Select format:",
                            options=[
                                "Copyable Format (Excel Paste Ready)",
                                "Research Format (Copy-Paste Ready)",
                                "Standard Format (Detailed)"
                            ],
                            help="Copyable: Tab-separated format matching single opposition copyable summary. Research: Matches your research Excel format. Standard: Detailed format with all data."
                        )

                        party_name_clean = st.session_state.party_name_searched.replace(' ', '_').replace(',', '')

                        if export_format == "Copyable Format (Excel Paste Ready)":
                            st.info("📋 **Copyable Format** - Matches single opposition copyable summary format. One row per opposition, tab-separated. Perfect for Excel paste!")
                            st.markdown("**Format:** `Marks | US GS | INT GS | Opp Start Date | Opp End Date | Result | TM Type | Serial | TM Type | Serial | ...`")

                            # Generate copyable format text
                            copyable_text = create_bulk_copyable_format(result)
                            copyable_excel_data = create_bulk_copyable_excel(result)

                            filename_copyable = f"{party_name_clean}_copyable_{result.get('search_date', '').replace(':', '').replace(' ', '_')}.xlsx"
                            filename_tsv = f"{party_name_clean}_copyable.tsv"

                            col1, col2 = st.columns(2)

                            with col1:
                                st.download_button(
                                    label="📥 Download Excel",
                                    data=copyable_excel_data,
                                    file_name=filename_copyable,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    type="primary"
                                )

                            with col2:
                                st.download_button(
                                    label="📋 Download as TSV (Tab-Separated)",
                                    data=copyable_text,
                                    file_name=filename_tsv,
                                    mime="text/tab-separated-values",
                                    use_container_width=True
                                )

                            # Show copyable text
                            with st.expander("📋 Copy Data for Excel Paste (Click to Expand)", expanded=False):
                                st.markdown("**Instructions:**")
                                st.markdown("1. Select all text below (Ctrl+A or Cmd+A)")
                                st.markdown("2. Copy (Ctrl+C or Cmd+C)")
                                st.markdown("3. Open Excel and click on the cell where you want to start pasting")
                                st.markdown("4. Paste (Ctrl+V or Cmd+V)")
                                st.markdown("5. Excel will automatically separate into columns!")
                                st.code(copyable_text, language=None)

                        elif export_format == "Research Format (Copy-Paste Ready)":
                            st.info("📋 **Research Format** - One row per opposition with alternating TM Type/Serial columns. Perfect for copy-pasting into your research Excel!")

                            # Optional metadata inputs
                            col1, col2 = st.columns(2)
                            with col1:
                                brand_id = st.text_input("Brand ID (Optional)", placeholder="e.g., 15011", help="Leave blank if unknown", key="brand_id_input")
                            with col2:
                                gvkey = st.text_input("GVKEY (Optional)", placeholder="e.g., 003007", help="Leave blank if unknown", key="gvkey_input")

                            # Generate research format Excel
                            research_excel_data = create_research_format_excel(result, brand_id=brand_id, gvkey=gvkey)
                            filename_research = f"{party_name_clean}_research_format_{result.get('search_date', '').replace(':', '').replace(' ', '_')}.xlsx"

                            col1, col2 = st.columns(2)

                            with col1:
                                st.download_button(
                                    label="📥 Download Research Excel",
                                    data=research_excel_data,
                                    file_name=filename_research,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    type="primary"
                                )

                            with col2:
                                # Generate copy-paste text
                                copyable_text = create_copyable_research_format(result, brand_id=brand_id, gvkey=gvkey)
                                st.download_button(
                                    label="📋 Download as TSV (Tab-Separated)",
                                    data=copyable_text,
                                    file_name=f"{party_name_clean}_research_copyable.tsv",
                                    mime="text/tab-separated-values",
                                    use_container_width=True
                                )

                            # Show copyable text
                            with st.expander("📋 Copy Data for Paste (Click to Expand)", expanded=False):
                                st.markdown("**Instructions:** Select all text below, copy (Ctrl+C or Cmd+C), then paste into Excel starting at your desired row.")
                                st.code(copyable_text, language=None)

                        else:
                            # Standard format
                            st.info("📊 **Standard Format** - Comprehensive format with all opposition data and metadata.")

                            # Generate standard Excel file
                            excel_data = create_bulk_party_excel(result)
                            filename = f"{party_name_clean}_bulk_oppositions_{result.get('search_date', '').replace(':', '').replace(' ', '_')}.xlsx"

                            col1, col2, col3 = st.columns([1, 1, 1])

                            with col1:
                                st.download_button(
                                    label="📥 Download Excel",
                                    data=excel_data,
                                    file_name=filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    type="primary"
                                )

                            with col2:
                                # JSON download
                                json_str = json.dumps(result, indent=2)
                                st.download_button(
                                    label="📥 Download JSON",
                                    data=json_str,
                                    file_name=f"{party_name_clean}_bulk_oppositions.json",
                                    mime="application/json",
                                    use_container_width=True
                                )
    
                        # Show errors if any
                        if stats.get('total_errors', 0) > 0:
                            with st.expander(f"⚠️ View Errors ({stats.get('total_errors', 0)} opposition(s) failed)", expanded=False):
                                error_rows = [opp for opp in result.get('oppositions', []) if opp.get('status') == 'FAILED']
                                for err_opp in error_rows:
                                    st.error(f"**Opposition {err_opp.get('opposition_number', '')}**: {err_opp.get('error_message', 'Unknown error')}")
    
                        # Detailed information
                        with st.expander("📋 View Complete Data", expanded=False):
                            # Show all oppositions in a table
                            all_data = []
                            for opp in result.get('oppositions', []):
                                all_data.append({
                                    'Opposition #': opp.get('opposition_number', ''),
                                    'Filing Date': opp.get('filing_date', ''),
                                    'Termination': opp.get('termination_date', ''),
                                    'Status': opp.get('status', ''),
                                    'Result': 'Sustained' if opp.get('result') == 1 else 'Dismissed' if opp.get('result') == 0 else 'Pending',
                                    'Plaintiff': opp.get('plaintiff', ''),
                                    'Defendant': opp.get('defendant', ''),
                                    'Serials': opp.get('serial_count', 0),
                                    'US Classes': opp.get('unique_us_classes', ''),
                                    'Intl Classes': opp.get('unique_international_classes', ''),
                                    'Cache Hit': opp.get('cache_hit_rate', ''),
                                    'Errors': opp.get('failed_serials', '')
                                })
    
                            df_all = pd.DataFrame(all_data)
                            st.dataframe(df_all, use_container_width=True, hide_index=True)

                    except Exception as e:
                        progress_bar.empty()
                        status_text.empty()
                        st.error(f"❌ Error during scraping: {str(e)}")
                        st.info("💡 **Suggestions:**\n- Check your internet connection\n- Try again in a few minutes\n- Verify the party name is correct")

    # TAB 3: Batch Company Validation
    with tab3:
        st.markdown("### Batch Company Validation")
        st.markdown("Upload an Excel file with company names to quickly check which companies have opposition data.")

        st.info("📋 **Purpose:** This tool performs a FAST check (page 1 only) to filter out companies without opposition data. If a company has ANY oppositions (even recent ones), it will be included in the results. This saves significant time before doing full scraping.")
        st.warning("⚡ **Quick Check Mode:** This checks only the first page of results for each company to maximize speed. If oppositions are found, the company is marked as having data.")

        # File upload
        uploaded_file = st.file_uploader(
            "Upload Excel file with company names",
            type=['xlsx', 'xls'],
            help="Upload an Excel file containing company names"
        )

        if uploaded_file:
            # Read Excel file
            try:
                df_input = pd.read_excel(uploaded_file)
                st.success(f"✅ Loaded {len(df_input)} rows from Excel file")

                # Show preview
                with st.expander("📋 Preview uploaded data", expanded=True):
                    st.dataframe(df_input.head(10), use_container_width=True)

                # Let user select column with company names
                st.markdown("### Select Company Name Column")
                column_name = st.selectbox(
                    "Which column contains company names?",
                    options=df_input.columns.tolist(),
                    help="Select the column that contains the company names to validate"
                )

                # Validation button
                st.markdown("---")
                validate_button = st.button("🔍 Start Validation", type="primary", use_container_width=True)

                if validate_button:
                    # Extract unique company names
                    company_names = df_input[column_name].dropna().unique().tolist()
                    total_companies = len(company_names)

                    st.info(f"🔍 Found {total_companies} unique companies to validate")

                    # Progress tracking
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    # Initialize scraper
                    scraper = USPTOOppositionScraper(API_KEY, CLAUDE_VISION_API_KEY, ANTHROPIC_API_KEY)

                    # Results storage
                    results = []

                    # Process each company
                    for idx, company_name in enumerate(company_names, 1):
                        status_text.text(f"Checking {idx}/{total_companies}: {company_name}")
                        progress_bar.progress(idx / total_companies)

                        try:
                            # Quick check for oppositions (optimized - only checks page 1)
                            check_result = scraper.quick_check_oppositions(
                                party_name=company_name,
                                max_retries=3
                            )

                            # Check for errors
                            if 'error' in check_result:
                                results.append({
                                    'Company Name': company_name,
                                    'Opposition Count': 0,
                                    'Has Data': 'Error',
                                    'Sample Dates': '',
                                    'Error': check_result.get('error', 'Unknown error')
                                })
                            else:
                                has_data = check_result.get('has_oppositions', False)
                                count = check_result.get('count', 0)
                                sample_dates = check_result.get('sample_dates', [])

                                # Format sample dates
                                dates_str = ', '.join(sample_dates) if sample_dates else ''

                                results.append({
                                    'Company Name': company_name,
                                    'Opposition Count': f"{count}+" if count >= 25 else str(count),  # 25 = page 1 limit
                                    'Has Data': 'Yes' if has_data else 'No',
                                    'Sample Dates': dates_str
                                })

                        except Exception as e:
                            results.append({
                                'Company Name': company_name,
                                'Opposition Count': 0,
                                'Has Data': 'Error',
                                'Sample Dates': '',
                                'Error': str(e)
                            })

                        # Increased delay to prevent connection resets (1.5 seconds between companies)
                        time.sleep(1.5)

                    # Clear progress indicators
                    progress_bar.empty()
                    status_text.empty()

                    # Create results dataframe
                    df_results = pd.DataFrame(results)

                    # Summary statistics
                    companies_with_data = len(df_results[df_results['Has Data'] == 'Yes'])
                    companies_without_data = len(df_results[df_results['Has Data'] == 'No'])
                    companies_error = len(df_results[df_results['Has Data'] == 'Error'])

                    st.success(f"✅ Validation Complete!")

                    # Show metrics
                    col1, col2, col3, col4 = st.columns(4)

                    with col1:
                        st.metric("Total Companies", total_companies)

                    with col2:
                        st.metric("With Oppositions", companies_with_data)

                    with col3:
                        st.metric("No Oppositions", companies_without_data)

                    with col4:
                        st.metric("Errors", companies_error)

                    # Show results table
                    st.markdown("---")
                    st.subheader("📊 Validation Results")

                    # Filter options
                    filter_option = st.radio(
                        "Show:",
                        options=["All Companies", "Only Companies WITH Oppositions", "Only Companies WITHOUT Oppositions"],
                        horizontal=True
                    )

                    if filter_option == "Only Companies WITH Oppositions":
                        df_display = df_results[df_results['Has Data'] == 'Yes']
                    elif filter_option == "Only Companies WITHOUT Oppositions":
                        df_display = df_results[df_results['Has Data'] == 'No']
                    else:
                        df_display = df_results

                    st.dataframe(df_display, use_container_width=True, hide_index=True)

                    # Download options
                    st.markdown("---")
                    st.subheader("💾 Download Results")

                    col1, col2 = st.columns(2)

                    with col1:
                        # Download all results
                        output_all = io.BytesIO()
                        with pd.ExcelWriter(output_all, engine='openpyxl') as writer:
                            df_results.to_excel(writer, sheet_name='All Companies', index=False)

                        st.download_button(
                            label="📥 Download All Results",
                            data=output_all.getvalue(),
                            file_name="company_validation_all.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

                    with col2:
                        # Download only companies with oppositions
                        df_with_data = df_results[df_results['Has Data'] == 'Yes']
                        output_valid = io.BytesIO()
                        with pd.ExcelWriter(output_valid, engine='openpyxl') as writer:
                            df_with_data.to_excel(writer, sheet_name='Companies With Data', index=False)

                        st.download_button(
                            label="📥 Download Companies WITH Oppositions",
                            data=output_valid.getvalue(),
                            file_name="companies_with_oppositions.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )

            except Exception as e:
                st.error(f"❌ Error reading Excel file: {str(e)}")
                st.info("💡 Make sure the file is a valid Excel file (.xlsx or .xls)")


    # Footer
    st.markdown("---")
    st.markdown("Built with Streamlit | Data from USPTO TSDR API")


if __name__ == "__main__":
    main()
